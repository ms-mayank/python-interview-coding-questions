from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import openpyxl
from datetime import datetime
import time
driver = webdriver.Edge()
driver.implicitly_wait(0.5)
driver.maximize_window()

# open excel sheet
sheetpath = "Excel/DMPmarkets.xlsx"
wb = openpyxl.load_workbook(sheetpath)
i=0 
while(i<4):
    wb.active = i
    sheet = wb.active
    print(sheet)
    driver.get("https://pfizer.sharepoint.com/sites/DMPProd/SitePages/DMP-Asset.aspx")
    # providing the total no of records
    for row in range(2, sheet.max_row+1):
        try:
            start_time = time.time()
            # Accessing all the data from Excel
            ExlAssetId = sheet.cell(row, 2).value
            ExlAssetName = sheet.cell(row, 3).value
            ExlSCO = sheet.cell(row, 4).value
            ExlClientPartner = sheet.cell(row, 5).value
            ExlBusinessOwner = sheet.cell(row, 6).value
            ExlTimeStamp = sheet.cell(row, 7)
            ExlStatus = sheet.cell(row, 8)
            ExlRemarks = sheet.cell(row, 9)
            ExlBrand = sheet.cell(row, 10).value # -> Change the relevant col no of Brand
            ExlBusinessUnit = sheet.cell(row, 11).value # -> Change the relevant col no of BU

            driver.get("https://pfizer.sharepoint.com/sites/DMPProd/SitePages/DMP-Asset.aspx")
            driver.get("https://pfizer.sharepoint.com/sites/DMPProd/SitePages/DMP-Asset.aspx#/editasset?"+str(ExlAssetId))
            time.sleep(2)
            '''
            # Finding Assetname field, clearing already present data and updating
            AssetName = driver.find_element(By.XPATH, '//*[@id="assetname"]')
            AssetName.clear()
            AssetName.send_keys(ExlAssetName)
            # AssetName.click()
            # time.sleep(2)
            # AssetName.send_keys(Keys.CONTROL,'a')
            # AssetName.send_keys(Keys.DELETE)
            # AssetName.send_keys(ExlAssetName)
            time.sleep(1)
            '''

            # Brand Name update
            Brand = driver.find_element(By.XPATH,'/html/body/div/div[2]/div[2]/div[2]/div[3]/section/article/div[1]/div/div/div[1]/div/div/div/div/div/div/div/div/div/div/div/div/div/div/div[2]/div[2]/div/form/fieldset[2]/div[1]/div/div/div[1]/div[2]/div/input')
            Brand.send_keys(Keys.DELETE)
            Brand.send_keys(Keys.DELETE)
            Brand.send_keys(Keys.DELETE)
            Brand.send_keys(Keys.DELETE)
            Brand.send_keys(ExlBrand)
            time.sleep(4)
            Brand.send_keys(Keys.ENTER)
            time.sleep(0.5)

            # Business Unit Update
            BusinessUnit = driver.find_element(By.XPATH,'/html/body/div/div[2]/div[2]/div[2]/div[3]/section/article/div[1]/div/div/div[1]/div/div/div/div/div/div/div/div/div/div/div/div/div/div/div[2]/div[2]/div/form/fieldset[2]/div[2]/div/div/div[1]/div[2]/div/input')
            BusinessUnit.send_keys(ExlBusinessUnit)
            time.sleep(2)
            BusinessUnit.send_keys(Keys.ENTER)
            time.sleep(0.5)
            '''
            # Finding SCO field, and updating
            SCO = driver.find_element(By.XPATH, '//*[@id="react-select-7-input"]')
            SCO.send_keys(ExlSCO)
            time.sleep(2)
            SCO.send_keys(Keys.ENTER)
            time.sleep(2)

            # Finding clientpartner field, and updating
            ClientPartner = driver.find_element(By.XPATH, '//*[@id="react-select-8-input"]')
            ClientPartner.send_keys(ExlClientPartner)
            time.sleep(2)
            ClientPartner.send_keys(Keys.ENTER)
            # print(Keys.ENTER)
            time.sleep(2)

            try:
                # Checking if Business Owner is selected or if selected then removed it.
                BusinessOwnerRmvbtn = driver.find_element(By.XPATH,'/html/body/div/div[2]/div[2]/div[2]/div[3]/section/article/div[1]/div/div/div[1]/div/div/div/div/div/div/div/div/div/div/div/div/div/div/div[2]/div[2]/div/form/fieldset[2]/div[8]/div/div/div[1]/div/div/div/div/span/div/button')
                BusinessOwnerRmvbtn.click()
                time.sleep(1)
                BusinessOwner = driver.find_element(By.XPATH,'/html/body/div[1]/div[2]/div[2]/div[2]/div[3]/section/article/div[1]/div/div/div[1]/div/div/div/div/div/div/div/div/div/div/div/div/div/div/div[2]/div[2]/div/form/fieldset[2]/div[8]/div[1]/div/div[1]/div/div/div/div/input')
            except:
                # else find the Business owner field
                BusinessOwner = driver.find_element(By.XPATH,'/html/body/div[1]/div[2]/div[2]/div[2]/div[3]/section/article/div[1]/div/div/div[1]/div/div/div/div/div/div/div/div/div/div/div/div/div/div/div[2]/div[2]/div/form/fieldset[2]/div[8]/div[1]/div/div[1]/div/div/div/div/input')
            finally:
                BusinessOwner.send_keys(ExlBusinessOwner)
                time.sleep(2)
                BusinessOwner.send_keys(Keys.ENTER)
            '''


            # Submit the changes           
            Submitbtn = driver.find_element(By.XPATH, '//*[@id="5af647a7-45f0-4a08-bf95-18104a86776a"]/div/div/div/div/div/div[1]/div/div[1]/div[2]/div[1]/button')
            Submitbtn.click()
            time.sleep(2)

            # Comment in pop up
            comment = driver.find_element(By.XPATH, '/html/body/div[2]/div/div/div/div[2]/div[2]/div/div/div[2]/div[2]/textarea')
            comment.click()
            comment.send_keys("Updated the record")
            time.sleep(2)
            

            # final save
            SaveUpdateBtn = driver.find_element(By.XPATH, '/html/body/div[2]/div/div/div/div[2]/div[2]/div/div/div[2]/div[2]/div[2]/button')
            SaveUpdateBtn.click()
            time.sleep(5)
          
            # To update the result in Excel
            now = datetime.now()
            ExlTimeStamp.value = now
            ExlStatus.value = "Yes"

        except:

            now = datetime.now()
            ExlStatus.value = "No"
            ExlTimeStamp.value = now

        finally:
            # getting the total time taken to update the record and updating it in remark field
            ExlRemarks.value = time.time() - start_time
    i=i+1
wb.save("DMPmarketsoutput.xlsx")
driver.quit()


